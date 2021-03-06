import yaml
import os
import csv
from bunch import bunchify
from collections import defaultdict
import openpyxl

#to get information of home or away match
excel_sheet = openpyxl.load_workbook('K:\IUB Sem-4\Data mining- B565\Data Mining Project\ODI_grounds_edited_excel.xlsx')
sheet = excel_sheet.get_sheet_by_name('ODI_grounds_edited_csv')
venue_country_dict = defaultdict(dict)
city_country_dict = defaultdict(dict)
for i in range(2, 195):
    stadium_name = sheet.cell(row=i, column=2).value
    country = sheet.cell(row=i, column=4).value
    cities = sheet.cell(row=i, column=3).value
    cname = country.split('!')
    city_name = cities.split(',')
    venue_country_dict[stadium_name] = (cname[0].strip())
    for i in range(len(city_name)):
        city_country_dict[city_name[i].strip()] = (cname[0].strip())
#Path to the directory in which all the files exist
path = "K://IUB Sem-4//Data mining- B565//odis//"
file_names = os.listdir(path) #stores the file names (with extensions) in a list (file_names)

#Iterates over each file to get the binary values of the features and ignores the matches in
#which the match outcome is not there (if the match is cancelled (due to rain or other reasons)
for files in file_names:
    print("filename", files)
    fpath = "K://IUB Sem-4//Data mining- B565//odis//" + files
    with open(fpath, 'r') as stream:
        datamap = yaml.safe_load(stream)
    x = bunchify(datamap)
    if(x.info.outcome.has_key('winner')):
        team1_bat_first = 0
        team1_bat_second = 0
        team2_bat_first = 0
        team2_bat_second = 0
        first_bat_team = ""
        team1_winner = 0
        team1_runs = 0
        team2_runs = 0
        team1_abv300 = 0
        team1_less200 = 0
        team1_between200n300 = 0
        team2_abv300 = 0
        team2_less200 = 0
        team2_between200n300 = 0
        team1_wicket_count = 0
        team2_wicket_count = 0
        team1_wicket = 0
        team2_wicket = 0
        home_match = 0
        first_bat_toss_winner = 0
        features_list = []

        if(x['info']['toss']['winner'] == x.info.teams[0]):
            if(x['info']['toss']['decision'] == "bat"):
                team1_bat_first = 1
                team2_bat_second = 1
                first_bat_team = x.info.teams[0]
            else:
                team1_bat_second = 1
                team2_bat_first = 1
                first_bat_team = x.info.teams[1]
        elif(x['info']['toss']['winner'] == x.info.teams[1]):
            if(x['info']['toss']['decision'] == "bat"):
                team2_bat_first = 1
                team1_bat_second = 1
                first_bat_team = x.info.teams[1]
            else:
                team2_bat_second = 1
                team1_bat_first = 1
                first_bat_team = x.info.teams[0]
        if((team1_bat_first & team2_bat_second) & (x.info.teams[0] == x.innings[0]['1st innings']['team'])):
            for d in x.innings[0]['1st innings']['deliveries']:
                for k in d:
                    team1_runs = team1_runs + d[k]['runs']['total']
                    if d[k].has_key('wicket'):
                        team2_wicket_count = team2_wicket_count + 1
            for d in x.innings[1]['2nd innings']['deliveries']:
                for k in d:
                    team2_runs = team2_runs + d[k]['runs']['total']
                    if d[k].has_key('wicket'):
                        team1_wicket_count = team1_wicket_count + 1
        elif((team2_bat_first & team1_bat_second) & (x.info.teams[1] == x.innings[0]['1st innings']['team'])):
            for d in x.innings[0]['1st innings']['deliveries']:
                for k in d:
                    team2_runs = team2_runs + d[k]['runs']['total']
                    if d[k].has_key('wicket'):
                        team1_wicket_count = team1_wicket_count + 1
            for d in x.innings[1]['2nd innings']['deliveries']:
                for k in d:
                    team1_runs = team1_runs + d[k]['runs']['total']
                    if d[k].has_key('wicket'):
                        team2_wicket_count = team2_wicket_count + 1

        #to get the team1 and team2 runs
        if(team1_runs >= 300):
            team1_abv300 = 1
        elif(team1_runs <= 200):
            team1_less200 = 1
        elif(team1_runs >= 200 & team1_runs <= 300):
            team1_between200n300 = 1
        if(team2_runs >= 300):
            team2_abv300 = 1
        elif(team2_runs <= 200):
            team2_less200 = 1
        elif(team2_runs >= 200 & team2_runs <= 300):
            team2_between200n300 = 1

        #to find the toss winner
        if(first_bat_team == x['info']['toss']['winner']):
            first_bat_toss_winner = 1

        #to get the binary value of wickets taken by both teams
        if(team1_wicket_count >= 5):
            team1_wicket = 1
        elif(team2_wicket_count >= 5):
            team2_wicket = 1

        #to get the binary value of the winner (team batting first win or loose)
        if(first_bat_team == x.info.outcome.winner):
            team1_winner = 1
        else:
            team1_winner = 0

        if(x.info.has_key('neutral_venue')):
            home_match = 1
            print "Neutral: "
        else:
            if(x.info.has_key('city')):
                print "City name: ", x['info']['city']
                if(x['info']['city'] == "hyderabad") & (first_bat_team == "India"):
                     home_match = 1
                if(x['info']['city'] == "Sind") & (first_bat_team == "Pakistan"):
                    home_match = 1
                if(x['info']['city'] == "Skating and Curling Club") & (first_bat_team == "Canada"):
                    home_match = 1
                if(x['info']['city'] == "Dharmasala") & (first_bat_team == "India"):
                    home_match = 1
                if(city_country_dict.has_key(x['info']['city'])):
                    if(first_bat_team == city_country_dict[x['info']['city']]):
                        home_match = 1
                    print "cities: ", first_bat_team, ",", x['info']['city'], ",", city_country_dict[x['info']['city']], home_match
            else:
                print "Venue name: ", x['info']['venue']
                if(x['info']['venue'] == "Dubai Sports City Cricket Stadium"):
                    if(first_bat_team == "United Arab Emirates"):
                        home_match = 1
                if(venue_country_dict.has_key(x['info']['venue'])):
                    if(first_bat_team == venue_country_dict[x['info']['venue']]):
                        home_match = 1
                    print "venue: ", first_bat_team, ",", x['info']['venue'], ",", venue_country_dict[x['info']['venue']], home_match

        '''geolocator = Nominatim()
        #city = x['info']['venue'].replace(',', '')
        if(x.info.has_key('city')):
            location = geolocator.geocode(x['info']['city'], timeout=15)
            address = [x.strip() for x in location._address.split(',')]
            if(address[-1]== "United Kingdom"):
                if(address[2] == "England") & (first_bat_team == "England"):
                    home_match = 1
            elif(address[-1] == first_bat_team):
                home_match = 1
        else:
            location = geolocator.geocode(x['info']['venue'], timeout=15)
            address = [x.strip() for x in location._address.split(',')]
            if(address[-1]== "United Kingdom"):
                if(address[2] == "England") & (first_bat_team == "England"):
                    home_match = 1
            elif(address[-1] == first_bat_team):
                home_match = 1'''
        features_list.append(team1_abv300)
        features_list.append(team1_less200)
        features_list.append(team1_between200n300)
        features_list.append(team1_wicket)
        features_list.append(team2_abv300)
        features_list.append(team2_less200)
        features_list.append(team2_between200n300)
        features_list.append(team2_wicket)
        features_list.append(first_bat_toss_winner)
        features_list.append(team1_winner)
        features_list.append(home_match)
        out = csv.writer(open('K://IUB Sem-4//Data mining- B565//features_dataset.csv',"ab"), delimiter=',',quoting=csv.QUOTE_MINIMAL)
        out.writerow(features_list)
